import openpyxl
import matplotlib.pyplot as plt
import statistics
import pandas as pd
import numpy as np
import random
import math
import 函数调用4 as functions
import sys

"""
说明：我不可能把开仓点设的过于严厉，但这样子的话，我怎么控制它在刚刚破线的时候不乱交易呢？是不是可以通过读取历史最大值来调整呢？
一开始没有可以随便交易，能解决一开始的跟仓问题，但这样会不会积重难返？
还有在高低位的高频的反复振荡有点危险
"""

workbook = openpyxl.load_workbook('数据表格2.xlsx')
sheet = workbook['测试用数据']
#记得在调用的.py也要改
"""
workbook = openpyxl.load_workbook('数据表格.xlsx')
sheet = workbook['整理后数据']

workbook = openpyxl.load_workbook('数据表格2.xlsx')
sheet = workbook['测试用数据']
"""

"""
记得下面的一系列东西也要在函数调用和测试里面改
"""
#这里是交易参数，可能后面还会补充开盘价，交易量等
trade_day = [cell.value for cell in sheet['A']][1:]
AU = [cell.value for cell in sheet['B']][1:]
AU_main = [cell.value for cell in sheet['D']][1:]
NAU = [cell.value for cell in sheet['C']][1:]
NAU_main = [cell.value for cell in sheet['E']][1:]
RMB = [cell.value for cell in sheet['H']][1:]
time_judge = [cell.value for cell in sheet['G']][1:]


def Backtesting(AU, NAU, trade_day, time_judge, RMB, grid_interval): #也就是说，我要在无套利区间底下做空，上面做多

    length = len(AU)
    ratio = np.full(length, np.nan)
    RSI_dif = np.full(length, np.nan)
    RSI_dif_logchange = np.full(length, np.nan)

    A_info_num = np.full(length, np.nan)
    NA_info_num = np.full(length, np.nan)
    Info_dif = np.full(length, np.nan)

    expma_trend = np.full(length, np.nan)

    p_change = np.full(length, np.nan)
    Trade_N = np.full(length, np.nan)
    position_list = np.zeros(length)

    AU_rsi_values = functions.__calculate_rsi(AU) #长度都和AU一致，为1982项
    NAU_rsi_values = functions.__calculate_rsi(NAU)

    Aj = functions.calculate_j(AU) #长度都和AU一致，为1982项
    NAj = functions.calculate_j(NAU)

    upper_list, lower_list = functions.Upper_Lower(AU, RMB, time_judge)
    relative_position = functions.calculate_relative_position(AU,NAU,RMB) #计算了每个ratio比较于前面的相对位置，也就是说，越大的我们交易越多
    #特别强调的是，这个relative本身就是拿i-1算出来的，也就是说，它的第i项都是拿i-1之前的数据弄出来的。

    count_avg = 0

    for i, d in enumerate(trade_day):  # i是明天，只能有i-1，i-2之类的,那么一共1982项
        """
        在这个地方我要使用for循环制造一些list，这些list要和交易没有直接关系，但注意不要使用未来函数。怎么说吧，用前天的无所谓，只要不用明天的就行。
        """

        # 制造ratio
        ratio[i] = (NAU[i] * RMB[i] / 31.1035) / AU[i]

        """
        main_intervals = functions.find_main_oscillation_intervals(ratio[:i-2]) if i>2 else 0
        list_main_intervals = []
        if i>2:
            for interval in main_intervals:
                list_main_intervals.append(interval[0])
                list_main_intervals.append(interval[1])
        else:
            list_main_intervals = [0,0]
        """

        # 确定信号大小，这个要用i-2 ,这个信号是由金价本身推出来的，应该问题不大？
        A_info_num[i] = 0.5 * math.log(Aj[i - 2] + Aj[i - 1]+0.0000001) if i > 1 else 0  #这个0.0000001是用来防止前n项的0造成的数学错误的
        NA_info_num[i] = 0.5 * math.log(NAj[i - 2] + NAj[i - 1]+0.0000001) if i > 1 else 0

        #制造info_dif
        Info_dif[i] = 1+(A_info_num[i]-NA_info_num[i])

        #制造RSI_dif
        RSI_dif[i] = NAU_rsi_values[i]- AU_rsi_values[i]

        #制造RSI_dif_logchange,使得差值的绝对值越大，交易数目就越大
        if i > 0:
            RSI_dif_log = math.log(abs(RSI_dif[i - 1]))
            RSI_dif_logchange[i] = RSI_dif_log * math.copysign(1, RSI_dif[i - 1])
        else:
            RSI_dif_logchange[i] = 0

        # 计算ratio的expma
        expma_values = functions.calculate_expma(ratio)

        # 制造p_change
        if i>0:
            if RSI_dif_logchange[i] >= 1:
                p_change[i] = 0
            elif RSI_dif_logchange[i]  <= -1:
                p_change[i] = 0
            else:
                p_change[i] = (RSI_dif_logchange[i])
        else:
            p_change[i] = 0

        """
        下面是进入交易的准备工作,记得要i-1开始，上面有些可以用i是因为这个是用来创list的，不直接用来交易
        """

        ratio_value = ratio[i-1] if i > 0 else 1
        prev_ratio_value = ratio[i - 2] if i > 1 else 1 #注意这两个，i-1是现在，i是明天的预测交易

        AU_price = AU[i]
        NAU_price = NAU[i]

        #这些制造用i是没啥问题的，毕竟上面没有用来参与交易，除了ratio_value这两个和change

        """
        下面是交易逻辑
        """

        #expma_trend代表了ratio的上升下降趋势,只拿昨天的是不行的，我得加上几天
        expma_trend[i] = expma_values[i] - expma_values[i - 1] if i>0 else 0 # 计算出trend #加动量的影响有点大
        expma_sign = expma_trend[i-1]

        KDE_lower = 0.9822138898433043
        KDE_upper = 1.0049151694030605


        if (ratio[i-1]<min(KDE_lower,lower_list[i-1]) and ratio[i-2]>max(KDE_lower,lower_list[i-2])) \
                or (ratio[i-1]> max(KDE_upper,upper_list[i-1]) and ratio[i-2]< min(KDE_upper,upper_list[i-2])) \
                and position_list[i-1]== 0: #这里是检查一下，如果出去界限却没有开仓就补上
            if  (ratio[i-1]<KDE_lower and ratio[i-2]>KDE_lower and ratio[i-3]>KDE_lower) :
                position_list[i] = 1
            elif (ratio[i-1]>KDE_upper and ratio[i-2]<KDE_upper and ratio[i-3]<KDE_upper) :
                position_list[i] = -1
            else:
                position_list[i] = position_list[i-1]
            count_avg += 1
            Trade_N[i] = position_list[i] - position_list[i - 1]
            continue

        #下面这个KDE加上0.01即可上1.3
        if ratio_value >KDE_upper+0.01 and (abs(position_list[i-1])!= 1): #1.005是拿的核密度算的，0.978也是，是密度=0.3,加0.01是创造一个利润带
            count_avg = 0
            if ratio_value > (1 + (1.1304 / AU_price)) if time_judge[i] == 1 else 1 + (
                    1.0456 / AU_price):  # 向上超出无套利区间（一定要写的原因是后面KDE可能会变）(减去了重复部分)
                if i > 150 and abs(relative_position[i - 1]) > 0.995 :  # 统一的150天是因为要等数据，这里的优先级比按照斜率交易高，呃呃，这一段也不交易
                    if expma_sign > 0.0003 and expma_trend[i - 3] > 0 and relative_position[i - 1] < 0:
                        position_list[i] = 1
                    elif expma_sign < -0.0003 and expma_trend[i - 3] < 0 and relative_position[i - 1] > 0:
                        position_list[i] = -1
                    else:
                        position_list[i] = position_list[i - 1]
                    Trade_N[i] = position_list[i] - position_list[i - 1]
                    continue

                # 我们可以直接抄底，分清是左侧还是右侧 （必要的）
                if expma_sign > 0.0005 and expma_sign < 0.0008 and relative_position[i - 1] > 0.95:  # 判断为左侧升
                    position_list[i] = -1

                elif expma_sign < -0.0005 and expma_sign > -0.0008 and relative_position[i - 1] < -0.95:  # 判断为右侧降
                    position_list[i] = 1

                else:
                    position_list[i] = position_list[i - 1]  # 不交易
            else:
                position_list[i] = position_list[i - 1]  # 不交易

        elif ratio_value < KDE_lower-0.01 and (abs(position_list[i-1])!= 1): #这句话的意思就是如果有了仓位咱就一直不交易
            count_avg = 0
            if ratio_value < 1 - (0.74 / AU_price + 0.49 * RMB[i] / (31.1035 * AU_price)): #虽说这个没必要但还是留着

                if i>150 and abs(relative_position[i-1])>0.995 \
                        and abs(expma_sign)>0.0003: #呃呃，这一段也不交易
                    if expma_sign > 0.0003 and expma_trend[i-3]>0 and relative_position[i-1] < 0:
                        position_list[i] = 1
                    elif expma_sign < -0.0003 and expma_trend[i-3]<0 and relative_position[i-1] > 0:
                        position_list[i] = -1
                    else:
                        position_list[i] = position_list[i - 1]
                    Trade_N[i] = position_list[i] - position_list[i - 1]
                    continue

                if expma_sign > 0.0005 and expma_sign<0.0008 and relative_position[i-1]<-0.95:  # 判断为左侧降
                    position_list[i] = 1

                elif expma_sign<-0.0005 and expma_sign>-0.0008 and relative_position[i-1]>0.95:  # 判断为右侧升
                    position_list[i] = -1

                else:
                    position_list[i] = position_list[i - 1]
            else:
                position_list[i] = position_list[i-1]


        elif i<20 and abs(relative_position[i-1])>0.99 \
                and (abs(position_list[i-1])!= 1 ): #这里是处理开头的突然ratio变化
            count_avg += 1
            if expma_sign > 0 and relative_position[i - 1] > 0.95:
                position_list[i] = 1
            elif expma_sign < 0 and relative_position[i - 1] < -0.95:
                position_list[i] = -1
            Trade_N[i] = position_list[i] - position_list[i - 1]
            continue

        elif i > 100 and count_avg == 100:  # 如果前一百个都在KDE线之间，我们大致认为它是振荡的，于是放宽条件。这玩意还是工作的。可以删掉
            count_avg += 1
            if relative_position[i - 1] > 0.5 and relative_position[i - 2] > 0.5:
                position_list[i] = -1
                Trade_N[i] = position_list[i] - position_list[i - 1]
                continue
            elif relative_position[i - 1] < -0.5 and relative_position[i - 2] < -0.5:
                position_list[i] = 1
                Trade_N[i] = position_list[i] - position_list[i - 1]
                continue
            else:
                position_list[i] = position_list[i - 1]

        elif ratio_value < upper_list[i - 1]  and ratio_value > lower_list[i - 1]:  # 在无套利区间里面，该平仓了,这个必须是最后的
            position_list[i] = 0
            count_avg += 1

        elif ratio_value > KDE_lower and ratio_value <KDE_upper: #不行就赶紧删掉
            position_list[i] = 0
            count_avg += 1

        else:  #不到无套利区间，又不在特殊值外面。或者，已经开完仓了，等等再说。
            position_list[i] = position_list[i - 1]
            count_avg += 1


        """
                elif ratio_value > 0.989 and ratio_value < 0.991 and position_list[i - 1] != 0: #加一段看看能不能在内部也交易
            #这两个数是threshold_density=0.8
            #如果不开仓平仓没有意义
            if abs(ratio[i-1]-ratio[i-10])>0.01:
                position_list[i] = 0
        """

        Trade_N[i] = position_list[i] - position_list[i - 1]  # 挪到这里来了，删掉了很多重复片段

        """
        这玩意得好好改改
        """
        if i > 100 and ratio_value <KDE_upper and ratio_value>KDE_lower\
                and abs(ratio[i - 1] - ratio[i - 5]) > 0.01 \
                and position_list[i - 1] == 0 \
                and (abs(relative_position[i - 1]) < 0.8):  # 这个小于0.8是用来看比较小的地方，整个if不触发
            count_avg += 1
            if relative_position[i - 1] > 0.5:
                if ratio[i - 1] - ratio[i - 5] < 0 and ratio[i - 1] - ratio[i - 10] < 0:  # 为什么这个不触发？？？
                    position_list[i] = position_list[i-1]
                    Trade_N[i] = position_list[i] - position_list[i - 1] #这里好像不触发
                    continue
                position_list[i] = -1
            elif relative_position[i - 1] < -0.5:
                if ratio[i - 1] - ratio[i - 5] > 0 and ratio[i - 1] - ratio[i - 10] > 0:
                    position_list[i] = position_list[i-1]
                    Trade_N[i] = position_list[i] - position_list[i - 1]
                    continue
                position_list[i] = 1

            Trade_N[i] = position_list[i] - position_list[i - 1]


        """
        elif i>100 and count_avg==100 :#如果前一百个都在KDE线之间，我们大致认为它是振荡的，于是放宽条件。这玩意还是工作的。
            count_avg += 1
            if relative_position[i - 1] > 0.5 and relative_position[i-2] > 0.5 :
                position_list[i] = -1
                Trade_N[i] = position_list[i] - position_list[i - 1]
                continue
            elif relative_position[i - 1] < -0.5 and relative_position[i-2]<-0.5 :
                position_list[i] = 1
                Trade_N[i] = position_list[i] - position_list[i - 1]
                continue
            else:
                position_list[i] = position_list[i-1]
        """

    data = {
        'trade_day': trade_day,
        'AU': AU,
        'NAU': NAU,
        'ratio': ratio,
        'Trade_N': Trade_N,
        'position_list': position_list,
    }
    df = pd.DataFrame(data) #导出csv文件
    df.to_excel('test_output.xlsx', index=False)

    return Trade_N, position_list, ratio ,df

Small_Trade_N,Small_postion_list,ratio,df = Backtesting(AU, NAU, trade_day, time_judge, RMB, grid_interval = 0.0012) #这个0.0012是拿训练集算出来的

Trade_N = []
for i in range(len(Small_Trade_N)):
    Trade_N.append((Small_Trade_N[i]))

postion_list = []
for i in range(len(Small_postion_list)):
    postion_list.append((Small_postion_list[i]))


#Trade_N,postion_list = Backtesting(AU,NAU,ratio,time_judge,RMB,p_change)
#print("交易向量是: {0}".format(Trade_N)) #长度均为n-1
#print("仓位情况是: {0}".format(postion_list[:-1]))
#print(len(AU))
#print(len(Trade_N))
#print(len(postion_list))
#print("交易和仓位长度是否一致（0为一致）: {0}".format(len(Trade_N)-len(postion_list))) #长度一致
#postion_list = postion_list[:-1]
#print(len(ratio)) #sum用来删不需要的分支
#print(df) #只要这个df不报错，就没有未来函数

"""
upper_list, lower_list = functions.Upper_Lower(AU, RMB, time_judge)

# 计算上界的最高值和平均值
upper_max = max(upper_list)

# 计算下界的最高值和平均值
lower_min = min(lower_list)

print(upper_max)
print(lower_min)
"""
third_quartile = np.percentile(ratio, 75)
#print(third_quartile)


Nega_Trade = 0
Posi_Trade = 0
No_Trade = 0
for i in range(len(Trade_N)):
    if Trade_N[i]>0:
        Posi_Trade+=1
    elif Trade_N[i]<0:
        Nega_Trade+=1
    else:
        No_Trade+=1
#print(Posi_Trade,Nega_Trade,No_Trade)



"""
上面的部分是交易的基本逻辑
"""

#[i-1]才是交易日？

def AU_Yield(AU,AU_main,NAU_main,postion_list):#这个是沪金收益率，记住是n-1项
    AU_Yield = []
    for i in range(1,len(AU)):
        if i==1:
            AU_Yield.append(0)
            continue
        if AU_main[i - 2] != AU_main[i-1] or NAU_main[i - 2] != NAU_main[i-1]:
            AU_Yield.append(0)
        else:
            temp_yield = round(((AU[i-1] / AU[i - 2]) - 1)*postion_list[i-2],9)
            AU_Yield.append(temp_yield)
    return AU_Yield
AU_Y = AU_Yield(AU,AU_main,NAU_main,postion_list)
#print("沪金收益率是: {0}".format(AU_Y))

def NAU_Yield(RMB,NAU,NAU_main,AU_main,postion_list):#这个是纽约金收益率，记住是n-1项
    NAU_Yield=[]
    for i in range(1,len(NAU)):
        if i==1:
            NAU_Yield.append(0)
            continue
        if NAU_main[i-2]!=NAU_main[i-1] or AU_main[i - 2] != AU_main[i-1]:
            NAU_Yield.append(0)
        else:
            temp_yield = round((((NAU[i-1]*RMB[i-1]) / (NAU[i - 2]*RMB[i-2])) - 1)*postion_list[i-2],9)
            NAU_Yield.append(temp_yield)
    return NAU_Yield
NAU_Y=NAU_Yield(RMB,NAU,NAU_main,AU_main,postion_list)
#print("纽约金收益率是: {0}".format(NAU_Y))

def final_Yield(AU_Y,NAU_Y):
    final_Y = []
    for i in range(len(AU_Y)):
        final_Y.append(NAU_Y[i]-AU_Y[i])
    return final_Y
final_Y = final_Yield(AU_Y,NAU_Y)

#print("综合收益率是: {0}".format(final_Y))
#print(len(final_Y))
#print(len(AU_Y))
#print(len(AU))

def Cumulative_Yield(final_Y):
    Cumulative_Yield = [final_Y[0]]
    for i in range(1,len(final_Y)):
        cumulative_return = Cumulative_Yield [i - 1] + final_Y[i]
        Cumulative_Yield.append(cumulative_return)
    return Cumulative_Yield
Cumulative_Y = Cumulative_Yield(final_Y)
#print("Cumulative_Yield is: {0}".format(Cumulative_Y))
#print(len(Cumulative_Y))



"""
上面的部分是各种收益率的计算
"""


drawdown_rate,max_dd,max_dd_duration, j, i = functions.max_drawdown(Cumulative_Y)
print("max_drawdown is: {0}".format(max_dd))
print("max_drawdown duration is: {0}".format(max_dd_duration))
print("最大回撤起始点：{0};最大回撤终止点{1}".format(j,i))
#print((Cumulative_Y[-1] / len(AU_Y)) * 365)


def Sharpe_Ratio(final_Y):
    growth_diff = []
    for i in range(len(final_Y)):
        growth_diff.append(final_Y[i])
    day_fluc = statistics.stdev(growth_diff)  # 日波动率
    aver_growth = statistics.mean(growth_diff)
    Sharpe_Ratio = (aver_growth * 15.5) / day_fluc  # 夏普指数
    return Sharpe_Ratio
Sharpe_Ratio=Sharpe_Ratio(final_Y)

def Sortino_Ratio(final_Y):
    growth_diff = []
    for i in range(len(final_Y)):
        growth_diff.append(final_Y[i])
    downside_diff = []  # 用以计算下行标准差
    for i in range(len(final_Y)):
        if final_Y[i] < 0:
            downside_diff.append(final_Y[i])
    std_downside = statistics.stdev(downside_diff) #下行标准差
    aver_growth = statistics.mean(growth_diff)
    Sortino_Ratio = (aver_growth * 15.5) / std_downside  # 索提诺指数
    return Sortino_Ratio
Sortino_Ratio = Sortino_Ratio(final_Y)

def Calmar_Ratio(Cumulative_Y,AU_Y,max_dd):
    anner_return_rate = (Cumulative_Y[-1] / len(AU_Y)) * 365
    Calmar_Ratio = anner_return_rate / max_dd
    return Calmar_Ratio
Calmar_Ratio = Calmar_Ratio(Cumulative_Y,AU_Y,max_dd)

def calculate_win_rate(Trade_N, Cumulative_Y, position_list): #没有完全解决
    total_trades = 0
    profitable_trades = 0
    prev_trade_idx = -1  # 上一次交易的索引

    for i in range(len(Trade_N)):
        if Trade_N[i] != 0:  # 排除不交易的情况
            if position_list[i] == 0:  # 判断是开仓还是平仓
                total_trades += 1
                if Cumulative_Y[i] >= Cumulative_Y[prev_trade_idx]:
                    profitable_trades += 1
            prev_trade_idx = i

    win_rate = (profitable_trades / total_trades) * 100
    return "{:.2f}%".format(win_rate)

win_rate = calculate_win_rate(Trade_N, Cumulative_Y,postion_list)

print("胜率是: {0}".format(win_rate))
print("Sharpe_Ratio is: {0}".format(Sharpe_Ratio))
print("Sortino_Ratio is: {0}".format(Sortino_Ratio))
print("Calmar_Ratio is: {0}".format(Calmar_Ratio))

"""
上面的部分是各种评测数据的计算
"""

"""
upper_list,lower_list = functions.Upper_Lower(AU,RMB,time_judge)
#这个是上下界的分布图，发现是在1.0004~1.00005，还有-0.9998~-0.9995
plt.figure(figsize=(13, 13))
font = {'family': 'Times New Roman',
        'weight': 'normal',
        'size': 15,
        }
data1 = upper_list[1:]
data2 = lower_list[1:]
plt.plot(data1, label="upper list")
plt.plot(data2, label="lower list")
plt.ylabel('value', fontsize=20)
plt.xlabel("time",labelpad=8.5, fontsize=20)
plt.legend(fontsize=20)
#plt.show()
"""


#这个是交易分布的散点图
#plt.scatter(range(len(AU)), Trade_N, marker="o", c="red" , s=0.5)
#plt.title("distribution of trade")
#plt.show()

#这个是累计收益率的图
#plt.style.use("dark_background")
#plt.figure(num=None, figsize=(12,6), frameon=True)
#plt.title("Cumulative Yield")
#plt.plot(range(len(AU_Y)), Cumulative_Y, color='green', marker='o', linewidth=1, markersize=0.5)
#plt.show()

#这个是Ratio的图,可以用来检验网格交易是否正确
#plt.style.use("dark_background")
#plt.figure(num=None, figsize=(12,6), frameon=True)
#plt.title("Ratio")
#plt.plot(range(len(AU_Y)+1), ratio, color='green', marker='o', linewidth=1, markersize=0.5)
#plt.show()


def plot_graph(ratio, AU, RMB, time_judge, Cumulative_Y, Trade_N, Nega_Trade, Posi_Trade, No_Trade,trade_day):
    """
    绘制图形函数

        参数：
        ratio：价格ratio列表
        AU：AU列表
        RMB：RMB列表
        time_judge：时间判断列表
        Cumulative_Y：Cumulative_Y列表
        Trade_N：Trade_N列表
        Nega_Trade：负交易次数
        Posi_Trade：正交易次数
        No_Trade：无交易次数
    """

    # 创建一个新的图形
    fig, ax1 = plt.subplots()

    # 设置第一个坐标轴
    ax1.plot(ratio, color='blue')
    ax1.set_xlabel('Time')
    ax1.set_ylabel('Ratio', color='blue')
    ax1.tick_params(axis='y', colors='blue')


    # 无套利区间
    upper, lower = functions.Upper_Lower(AU, RMB, time_judge)
    plt.plot(upper, label="upper list")
    plt.plot(lower, label="lower list")

    main_intervals = functions.find_main_oscillation_intervals(ratio)
    # 画主要振荡区间的虚线
    for interval in main_intervals:
        ax1.axhline(interval[0], linestyle='dashed', color='black')
        ax1.axhline(interval[1], linestyle='dashed', color='black')

    # 创建第二个坐标轴
    ax2 = ax1.twinx()
    ax2.plot(Cumulative_Y, color='red')
    ax2.set_xlabel('Time')
    ax2.set_ylabel('Cumulative Y', color='red')
    ax2.tick_params(axis='y', colors='red')


    # 初始化标注文本
    red_label = ''
    green_label = ''

    # 遍历Trade_N、ratio和Cumulative_Y的值
    for i in range(len(Trade_N)):
        if Trade_N[i] > 0:
            # 如果Trade_N的值大于零，则在ratio图像上标记一个红色点，大小和交易的绝对值成比例增大
            ax1.scatter(i, ratio[i], color='red', s=np.abs(Trade_N[i]) * 10)
            red_label = 'long NAU'
            # 在红色点上方标注Trade_N的值
            ax1.text(i, ratio[i], f'{Trade_N[i]:.3f}', color='red', ha='center', va='bottom')
        elif Trade_N[i] < 0:
            # 如果Trade_N的值小于零，则在ratio图像上标记一个绿色点，大小和交易的绝对值成比例增大
            ax1.scatter(i, ratio[i], color='green', s=np.abs(Trade_N[i]) * 10)
            green_label = 'long AU'
            # 在绿色点上方标注Trade_N的值
            ax1.text(i, ratio[i], f'{Trade_N[i]:.3f}', color='green', ha='center', va='top')

    # 添加标注文本

    ax1.text(0.5, 0.9, red_label, color='red', ha='center', va='center', transform=ax1.transAxes,
             bbox=dict(facecolor='white', edgecolor='red', boxstyle='round'))
    ax1.text(0.5, 0.1, green_label, color='green', ha='center', va='center', transform=ax1.transAxes,
             bbox=dict(facecolor='white', edgecolor='green', boxstyle='round'))

    ax1.set_xlabel('Time')



    # 创建表格
    Trade_Freq = round((Nega_Trade + Posi_Trade) / len(Trade_N), 3)
    table_data = [
        ['Neg. Trade', 'Pos. Trade', 'No Trade', 'Trade Frequency'],
        [Nega_Trade, Posi_Trade, No_Trade, Trade_Freq]
    ]

    table = ax2.table(cellText=table_data, colLabels=None, cellLoc='center', loc='center right',bbox=[0, -0.5, 1, 0.3])
    table.scale(0.6, 0.6)  # 缩小表格尺寸，可以根据需要调整缩放比例
    table.auto_set_font_size(False)
    table.set_fontsize(10)  # 调整表格字体大小

    # 调整图表布局
    fig.tight_layout()

    # 显示图形
    plt.show(block = True)


plot_graph(ratio, AU, RMB, time_judge, Cumulative_Y, Trade_N, Nega_Trade, Posi_Trade, No_Trade,trade_day)

"""
最后的这一部分是绘图
"""
sys.exit()
