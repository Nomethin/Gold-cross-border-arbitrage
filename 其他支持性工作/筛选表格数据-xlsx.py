import openpyxl
import matplotlib.pyplot as plt
import statistics
import pandas as pd
import numpy as np
import random
import math
import cntalib as talib
import 函数调用 as functions

# 打开原始 .xlsx 文件
workbook = openpyxl.load_workbook('沪深300指数.xlsx')  # 改这个东西的名字就可以了
sheet = workbook['沪深300指数']  # 这个也是

data = []
target_values = []

# 读取日期和值，并创建字典
for row in range(2, sheet.max_row + 1):
    date = sheet.cell(row=row, column=1).value
    value = sheet.cell(row=row, column=2).value
    data.append({'date': date, 'value': value})

# 读取目标日期
target_dates = []
for row in range(2, sheet.max_row + 1):
    target_date = sheet.cell(row=row, column=5).value
    target_dates.append(target_date)

# 根据目标日期筛选对应的值
for target_date in target_dates:
    for item in data:
        if item['date'] == target_date:
            target_values.append(item['value'])
            break
    else:
        target_values.append('')  # 如果没有找到对应的值，添加空白格

# 修改值为 42 的数据为前两个数据的平均值
for i in range(len(target_values)):
    if target_values[i] == 42:
        if i >= 2:
            average = (target_values[i-1] + target_values[i-2]) / 2
            target_values[i] = average

# 写入筛选出的值到第 F 列
sheet.cell(row=1, column=6, value='Value')  # 写入标题
for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=6, value=target_values[row - 2])

workbook.save('Modified_沪深300指数.xlsx')  # 记得改一下文件名
